from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.by import By
import openpyxl
from openpyxl import Workbook
from typing import Optional
from datetime import datetime
import pandas as pd
from openpyxl.styles import Font, numbers
import datetime as dt
from datetime import datetime, date

# Specify the filepath
filepath = "C:\\Users\\mtake\\Dropbox\\アンロックデータ\\unlock.xlsx"

def open_or_create_file(filename: str = filepath) -> Workbook:
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        workbook.create_sheet('project')
    return workbook

def update_or_insert(fd, prj: str, unCode: str, unVal: str, unDate: str, unTime: str):
    # Skip the entire row if unVal starts with $
    if unVal.startswith("$"):
        return

    sheet = fd['project']
    prj_col = 1
    unCode_col = 2
    unVal_col = 3
    unDate_col = 4
    unTime_col = 5
    row_to_update = None

    if unVal.endswith("%"):
        unVal = float(unVal[:-1]) / 100

    # Convert string to datetime
    unDate = datetime.strptime(unDate, "%d %b %Y").date()  # Get the date part only

    parsed_time = datetime.strptime(unTime, "%H:%M %p")
    if parsed_time.hour == 0 and parsed_time.minute == 0:
        unTime = "0:00"
    else:
        unTime = parsed_time.strftime("%H:%M")
        #unTime = parsed_time.strftime("%H:%M").lstrip("0")

    # Convert unTime to timedelta
    unTime = datetime.strptime(unTime, "%H:%M") - datetime.strptime("00:00", "%H:%M")

    for row in range(1, sheet.max_row + 1):
        existing_unDate = sheet.cell(row=row, column=unDate_col).value
        # If the unDate is of type datetime, extract only the date part
        if isinstance(existing_unDate, datetime):
            existing_unDate = existing_unDate.date()

        if sheet.cell(row=row, column=prj_col).value == prj and (unDate - existing_unDate).days < 10:
            row_to_update = row
            break

    if row_to_update is not None:
        sheet.cell(row=row_to_update, column=unCode_col).value = unCode
        sheet.cell(row=row_to_update, column=unVal_col).value = unVal
        sheet.cell(row=row_to_update, column=unDate_col).value = unDate
        sheet.cell(row=row_to_update, column=unTime_col).value = unTime
    else:
        sheet.append([prj, unCode, unVal, unDate, unTime])

    fd.save(filepath)

def update_colors(fd):
    sheet = fd['project']
    unDate_col = 4
    unVal_col = 3

    for row in range(2, sheet.max_row + 1):
        unDate = sheet.cell(row=row, column=unDate_col).value
        unVal = sheet.cell(row=row, column=unVal_col).value
        if isinstance(unDate, dt.datetime):
            unDate = unDate.date()
        if unVal >= 0.03:  # Check if unVal is greater or equal to 0.03
            sheet.cell(row=row, column=unVal_col).font = Font(color="FF0000")  # If so, change the color of this cell to red
        if (dt.datetime.now().date() - unDate).days > 10:
            sheet.cell(row=row, column=unDate_col).font = Font(color="FF0000")
        elif (dt.datetime.now().date() - unDate).days == 10:
            sheet.cell(row=row, column=unDate_col).font = Font(color="00FA9A")
        else:
            sheet.cell(row=row, column=unDate_col).font = Font(color="0000FF")

    fd.save(filepath)

def get_project_names():
    s = Service(ChromeDriverManager().install())
    driver = webdriver.Chrome(service=s)
    driver.get('https://token.unlocks.app/')
    project_elements = driver.find_elements(By.XPATH,
                                            '//*[@id="__next"]/div[1]/div[2]/div[2]/div[3]/div[2]/table/tbody/tr')

    fd = open_or_create_file()

    for idx, prj in enumerate(project_elements, start=1):
        unlockName = prj.find_element(By.XPATH, 'th/div/a/div/div[2]/p').text
        unlockCode = prj.find_element(By.XPATH, 'th/div/a/div/div[2]/div/span').text
        unlockPer = prj.find_element(By.XPATH, 'td[7]/a/div/div/div[1]/div[2]/p[1]').text
        unlockDate = prj.find_element(By.XPATH, 'td[7]/a/div/div/div[1]/div[6]/p[1]').text
        unlockTime = prj.find_element(By.XPATH, 'td[7]/a/div/div/div[1]/div[6]/p[2]').text

        print(f"{unlockName} : {unlockCode} : {unlockPer} : {unlockDate} : {unlockTime}")

        update_or_insert(fd, unlockName, unlockCode, unlockPer, unlockDate, unlockTime)

        if idx >= 20:
            break

    update_colors(fd)
    driver.quit()

if __name__ == '__main__':
    get_project_names()
